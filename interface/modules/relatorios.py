import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import sqlite3
import json
from datetime import datetime
from .base_module import BaseModule
from database import DB_NAME
from utils.formatters import format_date
from pdf_generators.relatorio_tecnico import gerar_pdf_relatorio

class RelatoriosModule(BaseModule):
    def setup_ui(self):
        # Container principal
        container = tk.Frame(self.frame, bg='#f8fafc')
        container.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header
        self.create_header(container)
        
        # Notebook para organizar seções
        self.notebook = ttk.Notebook(container)
        self.notebook.pack(fill="both", expand=True, pady=(20, 0))
        
        # Aba: Novo Relatório
        self.create_novo_relatorio_tab()
        
        # Aba: Lista de Relatórios
        self.create_lista_relatorios_tab()
        
        # Aba: Relatórios Gerais (nova funcionalidade)
        self.create_relatorios_gerais_tab()
        
        # Inicializar variáveis
        self.current_relatorio_id = None
        self.tecnicos_eventos = {}
        self.anexos_aba = {1: [], 2: [], 3: [], 4: []}
        
        # Carregar dados iniciais
        self.refresh_all_data()
        
    def create_header(self, parent):
        header_frame = tk.Frame(parent, bg='#f8fafc')
        header_frame.pack(fill="x", pady=(0, 20))
        
        title_label = tk.Label(header_frame, text="Relatórios Técnicos", 
                               font=('Arial', 18, 'bold'),
                               bg='#f8fafc',
                               fg='#1e293b')
        title_label.pack(side="left")
        
    def create_novo_relatorio_tab(self):
        # Frame da aba
        relatorio_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(relatorio_frame, text="Novo Relatório")
        
        # Scroll frame
        canvas = tk.Canvas(relatorio_frame, bg='white')
        scrollbar = ttk.Scrollbar(relatorio_frame, orient="vertical", command=canvas.yview)
        self.scrollable_relatorio = tk.Frame(canvas, bg='white')
        
        self.scrollable_relatorio.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.scrollable_relatorio, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Conteúdo do relatório
        self.create_relatorio_content(self.scrollable_relatorio)
        
    def create_relatorio_content(self, parent):
        # Frame principal dividido em duas colunas: conteúdo e indicadores
        main_content_frame = tk.Frame(parent, bg='white')
        main_content_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Coluna esquerda (conteúdo principal)
        left_frame = tk.Frame(main_content_frame, bg='white')
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        # Coluna direita (indicadores/dashboard) - largura aumentada
        right_frame = tk.Frame(main_content_frame, bg='#f8fafc', width=320)
        right_frame.pack(side="right", fill="y", padx=(10, 0))
        self.create_indicators_cards(right_frame)
        
        # Conteúdo principal
        self.create_cliente_section(left_frame)
        self.create_servico_section(left_frame)
        self.create_tecnicos_section(left_frame)
        self.create_equipamento_section(left_frame)
        self.create_vinculacao_section(left_frame)
        self.create_relatorio_buttons(left_frame)
        
    def create_cliente_section(self, parent):
        section_frame = self.create_section_frame(parent, "Identificação do Cliente")
        section_frame.pack(fill="x", pady=(0, 15))
        
        # Grid de campos
        fields_frame = tk.Frame(section_frame, bg='white')
        fields_frame.pack(fill="x")
        
        # Variáveis
        self.cliente_var = tk.StringVar()
        
        # Cliente com busca reativa
        tk.Label(fields_frame, text="Cliente *:", 
                 font=('Arial', 10, 'bold'), bg='white').grid(row=0, column=0, sticky="w", pady=5)
        
        cliente_frame = tk.Frame(fields_frame, bg='white')
        cliente_frame.grid(row=0, column=1, sticky="ew", padx=(10, 0), pady=5)
        
        self.cliente_combo = ttk.Combobox(cliente_frame, textvariable=self.cliente_var, width=40)
        self.cliente_combo.pack(side="left", fill="x", expand=True)
        
        # Botão para buscar/atualizar clientes
        refresh_clientes_btn = self.create_button(cliente_frame, "🔄", self.refresh_clientes, bg='#10b981')
        refresh_clientes_btn.pack(side="right", padx=(5, 0))
        
        # Configurar colunas
        fields_frame.grid_columnconfigure(1, weight=1)
        
    def create_servico_section(self, parent):
        section_frame = self.create_section_frame(parent, "Dados do Serviço")
        section_frame.pack(fill="x", pady=(0, 15))
        
        # Grid de campos
        fields_frame = tk.Frame(section_frame, bg='white')
        fields_frame.pack(fill="x")
        
        # Variáveis
        self.numero_relatorio_var = tk.StringVar()
        self.data_criacao_var = tk.StringVar(value=datetime.now().strftime('%d/%m/%Y'))
        self.formulario_servico_var = tk.StringVar()
        self.tipo_servico_var = tk.StringVar()
        self.data_recebimento_var = tk.StringVar()
        
        row = 0
        
        # Número do Relatório
        tk.Label(fields_frame, text="Número do Relatório:", 
                 font=('Arial', 10, 'bold'), bg='white').grid(row=row, column=0, sticky="w", pady=5)
        tk.Entry(fields_frame, textvariable=self.numero_relatorio_var, 
                 font=('Arial', 10), width=30).grid(row=row, column=1, sticky="ew", padx=(10, 0), pady=5)
        
        # Data de Criação
        tk.Label(fields_frame, text="Data de Criação:", 
                 font=('Arial', 10, 'bold'), bg='white').grid(row=row, column=2, sticky="w", pady=5, padx=(20, 0))
        tk.Entry(fields_frame, textvariable=self.data_criacao_var, 
                 font=('Arial', 10), width=15).grid(row=row, column=3, sticky="w", padx=(10, 0), pady=5)
        row += 1
        
        # Formulário de Serviço
        tk.Label(fields_frame, text="Formulário de Serviço:", 
                 font=('Arial', 10, 'bold'), bg='white').grid(row=row, column=0, sticky="w", pady=5)
        tk.Entry(fields_frame, textvariable=self.formulario_servico_var, 
                 font=('Arial', 10), width=30).grid(row=row, column=1, sticky="ew", padx=(10, 0), pady=5)
        
        # Tipo de Serviço
        tk.Label(fields_frame, text="Tipo de Serviço:", 
                 font=('Arial', 10, 'bold'), bg='white').grid(row=row, column=2, sticky="w", pady=5, padx=(20, 0))
        tipo_combo = ttk.Combobox(fields_frame, textvariable=self.tipo_servico_var, 
                                 values=["Manutenção", "Reparo", "Instalação", "Inspeção", "Consultoria"],
                                 width=12)
        tipo_combo.grid(row=row, column=3, sticky="w", padx=(10, 0), pady=5)
        row += 1
        
        # Data de Recebimento
        tk.Label(fields_frame, text="Data de Recebimento:", 
                 font=('Arial', 10, 'bold'), bg='white').grid(row=row, column=0, sticky="w", pady=5)
        tk.Entry(fields_frame, textvariable=self.data_recebimento_var, 
                 font=('Arial', 10), width=30).grid(row=row, column=1, sticky="ew", padx=(10, 0), pady=5)
        row += 1
        
        # Descrição do Serviço
        tk.Label(fields_frame, text="Descrição do Serviço:", 
                 font=('Arial', 10, 'bold'), bg='white').grid(row=row, column=0, sticky="nw", pady=5)
        self.descricao_text = scrolledtext.ScrolledText(fields_frame, height=3, width=40)
        self.descricao_text.grid(row=row, column=1, columnspan=3, sticky="ew", padx=(10, 0), pady=5)
        
        # Configurar colunas
        fields_frame.grid_columnconfigure(1, weight=1)
        
    def create_tecnicos_section(self, parent):
        section_frame = self.create_section_frame(parent, "Técnicos e Eventos")
        section_frame.pack(fill="x", pady=(0, 15))
        
        # Frame para adicionar técnico
        add_tecnico_frame = tk.Frame(section_frame, bg='white')
        add_tecnico_frame.pack(fill="x", pady=(0, 10))
        
        tk.Label(add_tecnico_frame, text="Adicionar Técnico:", 
                 font=('Arial', 10, 'bold'), bg='white').pack(side="left")
        
        self.tecnico_var = tk.StringVar()
        self.tecnico_combo = ttk.Combobox(add_tecnico_frame, textvariable=self.tecnico_var, width=30)
        self.tecnico_combo.pack(side="left", padx=(10, 0))
        
        # Botão para buscar/atualizar técnicos
        refresh_tecnicos_btn = self.create_button(add_tecnico_frame, "🔄", self.refresh_tecnicos, bg='#10b981')
        refresh_tecnicos_btn.pack(side="left", padx=(5, 0))
        
        add_tecnico_btn = self.create_button(add_tecnico_frame, "Adicionar Técnico", self.adicionar_tecnico)
        add_tecnico_btn.pack(side="left", padx=(10, 0))
        
        # Notebook para técnicos
        self.tecnicos_notebook = ttk.Notebook(section_frame)
        self.tecnicos_notebook.pack(fill="both", expand=True, pady=(10, 0))
        
    def create_equipamento_section(self, parent):
        section_frame = self.create_section_frame(parent, "Condição do Equipamento")
        section_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        # Notebook para as 4 abas
        self.equipamento_notebook = ttk.Notebook(section_frame)
        self.equipamento_notebook.pack(fill="both", expand=True, pady=(10, 0))
        
        # Aba 1: Condição Inicial
        self.create_aba1_condicao_inicial()
        
        # Aba 2: Peritagem do Subconjunto
        self.create_aba2_peritagem()
        
        # Aba 3: Desmembrando Unidade Compressora
        self.create_aba3_desmembrando()
        
        # Aba 4: Relação de Peças e Serviços
        self.create_aba4_pecas_servicos()
        
    def create_aba1_condicao_inicial(self):
        frame1 = tk.Frame(self.equipamento_notebook, bg='white')
        self.equipamento_notebook.add(frame1, text="1. Condição Inicial")
        
        # Título
        tk.Label(frame1, text="CONDIÇÃO ATUAL DO EQUIPAMENTO", 
                font=('Arial', 12, 'bold'), bg='white').pack(anchor="w", pady=5)
        
        # Campos da aba 1
        campos_aba1 = [
            "Cond. Encontrada",
            "Placa/N.Série",
            "Acoplamento",
            "Aspectos Rotores",
            "Válvulas Acopladas",
            "Data Recebimento"
        ]
        
        self.aba1_vars = {}
        fields_frame = tk.Frame(frame1, bg='white')
        fields_frame.pack(fill="x", padx=10, pady=5)
        
        for i, campo in enumerate(campos_aba1):
            row = i // 2
            col = (i % 2) * 2
            
            tk.Label(fields_frame, text=f"{campo}:", font=('Arial', 9, 'bold'), 
                    bg='white').grid(row=row, column=col, sticky="w", padx=5, pady=2)
            
            var = tk.StringVar()
            entry = tk.Entry(fields_frame, textvariable=var, font=('Arial', 9), width=25)
            entry.grid(row=row, column=col+1, sticky="ew", padx=5, pady=2)
            
            self.aba1_vars[campo] = var
        
        # Configurar grid
        for i in range(2):
            fields_frame.grid_columnconfigure(i*2+1, weight=1)
        
        # Seção de anexos
        self.create_anexos_section(frame1, 1)
        
    def create_aba2_peritagem(self):
        frame2 = tk.Frame(self.equipamento_notebook, bg='white')
        self.equipamento_notebook.add(frame2, text="2. Peritagem do Subconjunto")
        
        # Título
        tk.Label(frame2, text="DESACOPLANDO ELEMENTO COMPRESSOR DA CAIXA DE ACIONAMENTO", 
                font=('Arial', 12, 'bold'), bg='white').pack(anchor="w", pady=5)
        
        # Campos da aba 2
        campos_aba2 = [
            "Parafusos/Pinos",
            "Superfície Vedação",
            "Engrenagens",
            "Bico Injertor",
            "Rolamentos",
            "Aspecto Óleo",
            "Data"
        ]
        
        self.aba2_vars = {}
        fields_frame = tk.Frame(frame2, bg='white')
        fields_frame.pack(fill="x", padx=10, pady=5)
        
        for i, campo in enumerate(campos_aba2):
            row = i // 2
            col = (i % 2) * 2
            
            tk.Label(fields_frame, text=f"{campo}:", font=('Arial', 9, 'bold'), 
                    bg='white').grid(row=row, column=col, sticky="w", padx=5, pady=2)
            
            var = tk.StringVar()
            entry = tk.Entry(fields_frame, textvariable=var, font=('Arial', 9), width=25)
            entry.grid(row=row, column=col+1, sticky="ew", padx=5, pady=2)
            
            self.aba2_vars[campo] = var
        
        # Configurar grid
        for i in range(2):
            fields_frame.grid_columnconfigure(i*2+1, weight=1)
        
        # Seção de anexos
        self.create_anexos_section(frame2, 2)
        
    def create_aba3_desmembrando(self):
        frame3 = tk.Frame(self.equipamento_notebook, bg='white')
        self.equipamento_notebook.add(frame3, text="3. Desmembrando Unidade Compressora")
        
        # Título
        tk.Label(frame3, text="GRAU DE INTERFERÊNCIA NA DESMONTAGEM", 
                font=('Arial', 12, 'bold'), bg='white').pack(anchor="w", pady=5)
        
        # Campos da aba 3
        campos_aba3 = [
            "Interf. Desmontagem",
            "Aspecto Rotores",
            "Aspecto Carcaça",
            "Interf. Mancais",
            "Galeria Hidráulica",
            "Data Desmembração"
        ]
        
        self.aba3_vars = {}
        fields_frame = tk.Frame(frame3, bg='white')
        fields_frame.pack(fill="x", padx=10, pady=5)
        
        for i, campo in enumerate(campos_aba3):
            row = i // 2
            col = (i % 2) * 2
            
            tk.Label(fields_frame, text=f"{campo}:", font=('Arial', 9, 'bold'), 
                    bg='white').grid(row=row, column=col, sticky="w", padx=5, pady=2)
            
            var = tk.StringVar()
            entry = tk.Entry(fields_frame, textvariable=var, font=('Arial', 9), width=25)
            entry.grid(row=row, column=col+1, sticky="ew", padx=5, pady=2)
            
            self.aba3_vars[campo] = var
        
        # Configurar grid
        for i in range(2):
            fields_frame.grid_columnconfigure(i*2+1, weight=1)
        
        # Seção de anexos
        self.create_anexos_section(frame3, 3)
        
    def create_aba4_pecas_servicos(self):
        frame4 = tk.Frame(self.equipamento_notebook, bg='white')
        self.equipamento_notebook.add(frame4, text="4. Relação de Peças e Serviços")
        
        # Serviços Propostos
        tk.Label(frame4, text="SERVIÇOS PROPOSTO PARA REFORMA DO SUBCONJUNTO:", 
                font=('Arial', 10, 'bold'), bg='white').pack(anchor="w", pady=5)
        
        self.servicos_text = scrolledtext.ScrolledText(frame4, height=5, wrap=tk.WORD)
        self.servicos_text.pack(fill="x", padx=10, pady=2)
        
        # Peças Recomendadas
        tk.Label(frame4, text="PEÇAS RECOMENDADAS PARA REFORMA:", 
                font=('Arial', 10, 'bold'), bg='white').pack(anchor="w", pady=5)
        
        self.pecas_text = scrolledtext.ScrolledText(frame4, height=5, wrap=tk.WORD)
        self.pecas_text.pack(fill="x", padx=10, pady=2)
        
        # Data
        data_frame = tk.Frame(frame4, bg='white')
        data_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(data_frame, text="DATA:", font=('Arial', 10, 'bold'), bg='white').pack(side="left")
        self.data_pecas_var = tk.StringVar()
        tk.Entry(data_frame, textvariable=self.data_pecas_var, font=('Arial', 10), width=20).pack(side="left", padx=(10, 0))
        
        # Seção de anexos
        self.create_anexos_section(frame4, 4)
        
    def create_anexos_section(self, parent, aba_numero):
        # Título da seção de anexos
        anexos_frame = tk.LabelFrame(parent, text="Anexos", font=('Arial', 10, 'bold'), bg='white')
        anexos_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Frame para lista de anexos
        lista_anexos_frame = tk.Frame(anexos_frame, bg='white')
        lista_anexos_frame.pack(fill="both", expand=True)
        
        # Listbox para anexos
        anexos_listbox = tk.Listbox(lista_anexos_frame, height=4)
        anexos_listbox.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Armazenar referência
        setattr(self, f'anexos_listbox_aba{aba_numero}', anexos_listbox)
        
        # Botões
        btn_frame = tk.Frame(anexos_frame, bg='white')
        btn_frame.pack(fill="x", padx=5, pady=5)
        
        add_anexo_btn = self.create_button(btn_frame, "Adicionar Anexo", 
                                          lambda: self.adicionar_anexo(aba_numero), bg='#10b981')
        add_anexo_btn.pack(side="left", padx=(0, 5))
        
        remove_anexo_btn = self.create_button(btn_frame, "Remover Anexo", 
                                             lambda: self.remover_anexo(aba_numero), bg='#dc2626')
        remove_anexo_btn.pack(side="left")
        
    def create_vinculacao_section(self, parent):
        section_frame = self.create_section_frame(parent, "Vinculação com Cotação")
        section_frame.pack(fill="x", pady=(0, 15))
        
        # Campo para selecionar cotação
        vinc_frame = tk.Frame(section_frame, bg='white')
        vinc_frame.pack(fill="x")
        
        tk.Label(vinc_frame, text="Cotação Vinculada:", 
                 font=('Arial', 10, 'bold'), bg='white').pack(side="left")
        
        self.cotacao_var = tk.StringVar()
        self.cotacao_combo = ttk.Combobox(vinc_frame, textvariable=self.cotacao_var, width=40)
        self.cotacao_combo.pack(side="left", padx=(10, 0))
        
    def create_relatorio_buttons(self, parent):
        buttons_frame = tk.Frame(parent, bg='white')
        buttons_frame.pack(fill="x", pady=(20, 0))
        
        # Botões
        novo_btn = self.create_button(buttons_frame, "Novo Relatório", self.novo_relatorio, bg='#e2e8f0', fg='#475569')
        novo_btn.pack(side="left", padx=(0, 10))
        
        salvar_btn = self.create_button(buttons_frame, "Salvar Relatório", self.salvar_relatorio)
        salvar_btn.pack(side="left", padx=(0, 10))
        
        gerar_pdf_btn = self.create_button(buttons_frame, "Gerar PDF", self.gerar_pdf, bg='#10b981')
        gerar_pdf_btn.pack(side="right")
        
    def create_lista_relatorios_tab(self):
        # Frame da aba
        lista_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(lista_frame, text="Lista de Relatórios")
        
        # Container
        container = tk.Frame(lista_frame, bg='white', padx=20, pady=20)
        container.pack(fill="both", expand=True)
        
        # Frame de busca
        search_frame, self.search_var = self.create_search_frame(container, command=self.buscar_relatorios)
        search_frame.pack(fill="x", pady=(0, 15))
        
        # Treeview para lista
        columns = ("numero", "cliente", "data", "responsavel", "tipo")
        self.relatorios_tree = ttk.Treeview(container, columns=columns, show="headings", height=15)
        
        # Cabeçalhos
        self.relatorios_tree.heading("numero", text="Número")
        self.relatorios_tree.heading("cliente", text="Cliente")
        self.relatorios_tree.heading("data", text="Data")
        self.relatorios_tree.heading("responsavel", text="Responsável")
        self.relatorios_tree.heading("tipo", text="Tipo")
        
        # Larguras
        self.relatorios_tree.column("numero", width=150)
        self.relatorios_tree.column("cliente", width=200)
        self.relatorios_tree.column("data", width=100)
        self.relatorios_tree.column("responsavel", width=150)
        self.relatorios_tree.column("tipo", width=120)
        
        # Scrollbar
        lista_scrollbar = ttk.Scrollbar(container, orient="vertical", command=self.relatorios_tree.yview)
        self.relatorios_tree.configure(yscrollcommand=lista_scrollbar.set)
        
        # Pack
        self.relatorios_tree.pack(side="left", fill="both", expand=True)
        lista_scrollbar.pack(side="right", fill="y")
        
        # Botões da lista
        lista_buttons = tk.Frame(container, bg='white')
        lista_buttons.pack(fill="x", pady=(15, 0))
        
        editar_btn = self.create_button(lista_buttons, "Editar", self.editar_relatorio)
        editar_btn.pack(side="left", padx=(0, 10))
        
        duplicar_btn = self.create_button(lista_buttons, "Duplicar", self.duplicar_relatorio, bg='#f59e0b')
        duplicar_btn.pack(side="left", padx=(0, 10))
        
        gerar_pdf_lista_btn = self.create_button(lista_buttons, "Gerar PDF", self.gerar_pdf_selecionado, bg='#10b981')
        gerar_pdf_lista_btn.pack(side="right")
        
    def refresh_all_data(self):
        """Atualizar todos os dados do módulo"""
        self.refresh_clientes()
        self.refresh_tecnicos()
        self.refresh_cotacoes()
        self.carregar_relatorios()
        
    def refresh_clientes(self):
        """Atualizar lista de clientes"""
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        
        try:
            c.execute("SELECT id, nome FROM clientes ORDER BY nome")
            clientes = c.fetchall()
            
            self.clientes_dict = {f"{nome} (ID: {id})": id for id, nome in clientes}
            cliente_values = list(self.clientes_dict.keys())
            
            self.cliente_combo['values'] = cliente_values
            
            print(f"Clientes carregados no relatório: {len(cliente_values)}")  # Debug
            
        except sqlite3.Error as e:
            self.show_error(f"Erro ao carregar clientes: {e}")
        finally:
            conn.close()
            
    def refresh_tecnicos(self):
        """Atualizar lista de técnicos (agora baseado em usuários)"""
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        
        try:
            # Buscar usuários em vez de técnicos
            c.execute("SELECT id, nome_completo FROM usuarios WHERE nome_completo IS NOT NULL ORDER BY nome_completo")
            tecnicos = c.fetchall()
            
            self.tecnicos_dict = {f"{nome} (ID: {id})": id for id, nome in tecnicos}
            tecnico_values = list(self.tecnicos_dict.keys())
            
            self.tecnico_combo['values'] = tecnico_values
            
            print(f"Técnicos carregados: {len(tecnico_values)}")  # Debug
            
        except sqlite3.Error as e:
            self.show_error(f"Erro ao carregar técnicos: {e}")
        finally:
            conn.close()
            
    def refresh_cotacoes(self):
        """Atualizar lista de cotações"""
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        
        try:
            c.execute("SELECT id, numero_proposta FROM cotacoes ORDER BY numero_proposta")
            cotacoes = c.fetchall()
            
            self.cotacoes_dict = {f"{numero} (ID: {id})": id for id, numero in cotacoes}
            cotacao_values = [""] + list(self.cotacoes_dict.keys())  # Incluir opção vazia
            
            self.cotacao_combo['values'] = cotacao_values
            
        except sqlite3.Error as e:
            self.show_error(f"Erro ao carregar cotações: {e}")
        finally:
            conn.close()
            
    def adicionar_tecnico(self):
        """Adicionar técnico ao relatório"""
        tecnico_str = self.tecnico_var.get().strip()
        if not tecnico_str:
            self.show_warning("Selecione um técnico.")
            return
            
        tecnico_id = self.tecnicos_dict.get(tecnico_str)
        if not tecnico_id:
            self.show_warning("Técnico selecionado inválido.")
            return
            
        # Verificar se técnico já foi adicionado
        if tecnico_id in self.tecnicos_eventos:
            self.show_warning("Técnico já foi adicionado.")
            return
            
        # Criar aba para o técnico
        tecnico_nome = tecnico_str.split(' (ID:')[0]
        tecnico_frame = tk.Frame(self.tecnicos_notebook, bg='white')
        self.tecnicos_notebook.add(tecnico_frame, text=tecnico_nome)
        
        # Frame para adicionar eventos
        add_evento_frame = tk.Frame(tecnico_frame, bg='white')
        add_evento_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(add_evento_frame, text="Data/Hora:", font=('Arial', 9, 'bold'), bg='white').grid(row=0, column=0, sticky="w")
        data_hora_var = tk.StringVar(value=datetime.now().strftime('%d/%m/%Y %H:%M'))
        tk.Entry(add_evento_frame, textvariable=data_hora_var, width=15).grid(row=0, column=1, padx=5)
        
        tk.Label(add_evento_frame, text="Tipo:", font=('Arial', 9, 'bold'), bg='white').grid(row=0, column=2, sticky="w", padx=(10, 0))
        tipo_var = tk.StringVar(value="Início")
        tipo_combo = ttk.Combobox(add_evento_frame, textvariable=tipo_var, 
                                 values=["Início", "Fim", "Pausa"], width=10, state="readonly")
        tipo_combo.grid(row=0, column=3, padx=5)
        
        tk.Label(add_evento_frame, text="Evento:", font=('Arial', 9, 'bold'), bg='white').grid(row=1, column=0, sticky="w")
        evento_var = tk.StringVar()
        tk.Entry(add_evento_frame, textvariable=evento_var, width=50).grid(row=1, column=1, columnspan=2, padx=5, sticky="ew")
        
        add_evento_btn = self.create_button(add_evento_frame, "Adicionar", 
                                           lambda: self.adicionar_evento(tecnico_id, data_hora_var, tipo_var, evento_var, eventos_tree))
        add_evento_btn.grid(row=1, column=3, padx=5)
        
        # Grid de eventos
        eventos_tree = ttk.Treeview(tecnico_frame, columns=("data_hora", "tipo", "evento"), show="headings", height=8)
        eventos_tree.heading("data_hora", text="Data/Hora")
        eventos_tree.heading("tipo", text="Tipo")
        eventos_tree.heading("evento", text="Evento")
        eventos_tree.pack(fill="both", expand=True, padx=10, pady=5)
        
        # Botão para remover evento
        remove_evento_btn = self.create_button(tecnico_frame, "Remover Evento", 
                                              lambda: self.remover_evento(eventos_tree), bg='#dc2626')
        remove_evento_btn.pack(pady=5)
        
        # Armazenar referências
        self.tecnicos_eventos[tecnico_id] = {
            'nome': tecnico_nome,
            'tree': eventos_tree,
            'data_hora_var': data_hora_var,
            'tipo_var': tipo_var,
            'evento_var': evento_var
        }
        
        # Limpar seleção
        self.tecnico_var.set("")
        
    def adicionar_evento(self, tecnico_id, data_hora_var, tipo_var, evento_var, tree):
        """Adicionar evento ao técnico"""
        data_hora = data_hora_var.get().strip()
        tipo = tipo_var.get()
        evento = evento_var.get().strip()
        
        if not data_hora or not evento:
            self.show_warning("Preencha data/hora e evento.")
            return
            
        # Adicionar à lista
        tree.insert("", "end", values=(data_hora, tipo, evento))
        
        # Limpar campos
        evento_var.set("")
        data_hora_var.set(datetime.now().strftime('%d/%m/%Y %H:%M'))
        
    def remover_evento(self, tree):
        """Remover evento selecionado"""
        selected = tree.selection()
        if not selected:
            self.show_warning("Selecione um evento para remover.")
            return
            
        for item in selected:
            tree.delete(item)
            
    def adicionar_anexo(self, aba_numero):
        """Adicionar anexo à aba"""
        filepath = filedialog.askopenfilename(
            title=f"Selecionar Anexo para Aba {aba_numero}",
            filetypes=[("Todos os arquivos", "*.*"), ("Imagens", "*.jpg *.jpeg *.png"), ("PDFs", "*.pdf")]
        )
        
        if not filepath:
            return
            
        # Criar um dicionário com informações do anexo
        nome_arquivo = filepath.split('/')[-1]
        anexo_info = {
            'nome': nome_arquivo,
            'caminho': filepath,
            'descricao': f'Anexo da Aba {aba_numero}'
        }
        
        # Adicionar à lista de anexos
        self.anexos_aba[aba_numero].append(anexo_info)
        
        # Atualizar listbox
        listbox = getattr(self, f'anexos_listbox_aba{aba_numero}')
        listbox.insert(tk.END, nome_arquivo)  # Mostrar apenas o nome do arquivo
        
    def remover_anexo(self, aba_numero):
        """Remover anexo selecionado"""
        listbox = getattr(self, f'anexos_listbox_aba{aba_numero}')
        selected = listbox.curselection()
        
        if not selected:
            self.show_warning("Selecione um anexo para remover.")
            return
            
        index = selected[0]
        
        # Remover da lista
        self.anexos_aba[aba_numero].pop(index)
        
        # Remover da listbox
        listbox.delete(index)
        
    def _debug_anexos_json(self, aba_num):
        """Debug function para ver o que está sendo salvo nos anexos"""
        anexos = self.anexos_aba[aba_num] if self.anexos_aba[aba_num] else []
        json_result = json.dumps(anexos) if anexos else "[]"
        print(f"DEBUG: Aba {aba_num} tem {len(anexos)} anexos: {json_result}")
        return json_result
    
    def novo_relatorio(self):
        """Limpar formulário para novo relatório"""
        self.current_relatorio_id = None
        
        # Limpar campos básicos
        self.cliente_var.set("")
        self.numero_relatorio_var.set("")
        self.data_criacao_var.set(datetime.now().strftime('%d/%m/%Y'))
        self.formulario_servico_var.set("")
        self.tipo_servico_var.set("")
        self.data_recebimento_var.set("")
        self.descricao_text.delete("1.0", tk.END)
        
        # Limpar abas do equipamento
        for var_dict in [self.aba1_vars, self.aba2_vars, self.aba3_vars]:
            for var in var_dict.values():
                var.set("")
                
        self.servicos_text.delete("1.0", tk.END)
        self.pecas_text.delete("1.0", tk.END)
        self.data_pecas_var.set("")
        
        # Limpar técnicos
        for tab in self.tecnicos_notebook.tabs():
            self.tecnicos_notebook.forget(tab)
        self.tecnicos_eventos = {}
        
        # Limpar anexos
        for aba_num in range(1, 5):
            self.anexos_aba[aba_num] = []
            listbox = getattr(self, f'anexos_listbox_aba{aba_num}')
            listbox.delete(0, tk.END)
        
        # Limpar cotação
        self.cotacao_var.set("")
        
        # Gerar número automático
        numero = f"REL-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        self.numero_relatorio_var.set(numero)
        
    def limpar_formulario_edicao(self):
        """Limpar formulário para edição sem apagar anexos"""
        self.current_relatorio_id = None
        
        # Limpar campos básicos
        self.cliente_var.set("")
        self.numero_relatorio_var.set("")
        self.data_criacao_var.set("")
        self.formulario_servico_var.set("")
        self.tipo_servico_var.set("")
        self.data_recebimento_var.set("")
        self.descricao_text.delete("1.0", tk.END)
        
        # Limpar abas do equipamento
        for var_dict in [self.aba1_vars, self.aba2_vars, self.aba3_vars]:
            for var in var_dict.values():
                var.set("")
                
        self.servicos_text.delete("1.0", tk.END)
        self.pecas_text.delete("1.0", tk.END)
        self.data_pecas_var.set("")
        
        # Limpar técnicos
        for tab in self.tecnicos_notebook.tabs():
            self.tecnicos_notebook.forget(tab)
        self.tecnicos_eventos = {}
        
        # Limpar cotação
        self.cotacao_var.set("")
        
        # NÃO limpar anexos - eles serão carregados depois
        
    def salvar_relatorio(self):
        """Salvar relatório no banco de dados"""
        # Validações
        cliente_str = self.cliente_var.get().strip()
        numero = self.numero_relatorio_var.get().strip()
        
        if not cliente_str:
            self.show_warning("Selecione um cliente.")
            return
            
        if not numero:
            self.show_warning("Informe o número do relatório.")
            return
            
        # Obter ID do cliente
        cliente_id = self.clientes_dict.get(cliente_str)
        if not cliente_id:
            self.show_warning("Cliente selecionado inválido.")
            return
            
        # Obter ID da cotação (opcional)
        cotacao_str = self.cotacao_var.get().strip()
        cotacao_id = None
        if cotacao_str:
            cotacao_id = self.cotacoes_dict.get(cotacao_str)
            
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        
        try:
            # Preparar dados do relatório
            dados_relatorio = (
                numero,
                cliente_id,
                self.user_id,
                datetime.now().strftime('%Y-%m-%d') if not self.current_relatorio_id else None,
                self.formulario_servico_var.get().strip(),
                self.tipo_servico_var.get().strip(),
                self.descricao_text.get("1.0", tk.END).strip(),
                self.data_recebimento_var.get().strip(),
                
                # Aba 1
                self.aba1_vars.get("Cond. Encontrada", tk.StringVar()).get(),
                self.aba1_vars.get("Placa/N.Série", tk.StringVar()).get(),
                self.aba1_vars.get("Acoplamento", tk.StringVar()).get(),
                self.aba1_vars.get("Aspectos Rotores", tk.StringVar()).get(),
                self.aba1_vars.get("Válvulas Acopladas", tk.StringVar()).get(),
                self.aba1_vars.get("Data Recebimento", tk.StringVar()).get(),
                
                # Aba 2
                self.aba2_vars.get("Parafusos/Pinos", tk.StringVar()).get(),
                self.aba2_vars.get("Superfície Vedação", tk.StringVar()).get(),
                self.aba2_vars.get("Engrenagens", tk.StringVar()).get(),
                self.aba2_vars.get("Bico Injertor", tk.StringVar()).get(),
                self.aba2_vars.get("Rolamentos", tk.StringVar()).get(),
                self.aba2_vars.get("Aspecto Óleo", tk.StringVar()).get(),
                self.aba2_vars.get("Data", tk.StringVar()).get(),
                
                # Aba 3
                self.aba3_vars.get("Interf. Desmontagem", tk.StringVar()).get(),
                self.aba3_vars.get("Aspecto Rotores", tk.StringVar()).get(),
                self.aba3_vars.get("Aspecto Carcaça", tk.StringVar()).get(),
                self.aba3_vars.get("Interf. Mancais", tk.StringVar()).get(),
                self.aba3_vars.get("Galeria Hidráulica", tk.StringVar()).get(),
                self.aba3_vars.get("Data Desmembração", tk.StringVar()).get(),
                
                # Aba 4
                self.servicos_text.get("1.0", tk.END).strip(),
                self.pecas_text.get("1.0", tk.END).strip(),
                self.data_pecas_var.get().strip(),
                
                # Outros
                cotacao_id,
                "",  # tempo_trabalho_total
                "",  # tempo_deslocamento_total
                "",  # fotos
                self._debug_anexos_json(1),  # anexos_aba1
                self._debug_anexos_json(2),  # anexos_aba2
                self._debug_anexos_json(3),  # anexos_aba3
                self._debug_anexos_json(4)   # anexos_aba4
            )
            
            if self.current_relatorio_id:
                # Atualizar relatório existente
                c.execute("""
                    UPDATE relatorios_tecnicos SET
                        numero_relatorio = ?, cliente_id = ?, formulario_servico = ?,
                        tipo_servico = ?, descricao_servico = ?, data_recebimento = ?,
                        condicao_encontrada = ?, placa_identificacao = ?, acoplamento = ?,
                        aspectos_rotores = ?, valvulas_acopladas = ?, data_recebimento_equip = ?,
                        parafusos_pinos = ?, superficie_vedacao = ?, engrenagens = ?,
                        bico_injertor = ?, rolamentos = ?, aspecto_oleo = ?, data_peritagem = ?,
                        interf_desmontagem = ?, aspecto_rotores_aba3 = ?, aspecto_carcaca = ?,
                        interf_mancais = ?, galeria_hidraulica = ?, data_desmembracao = ?,
                        servicos_propostos = ?, pecas_recomendadas = ?, data_pecas = ?,
                        cotacao_id = ?, tempo_trabalho_total = ?, tempo_deslocamento_total = ?,
                        fotos = ?, anexos_aba1 = ?, anexos_aba2 = ?, anexos_aba3 = ?, anexos_aba4 = ?
                    WHERE id = ?
                """, (dados_relatorio[0], dados_relatorio[1]) + dados_relatorio[4:] + (self.current_relatorio_id,))
                
                # Remover eventos antigos
                c.execute("DELETE FROM eventos_campo WHERE relatorio_id = ?", (self.current_relatorio_id,))
                relatorio_id = self.current_relatorio_id
            else:
                # Inserir novo relatório
                c.execute("""
                    INSERT INTO relatorios_tecnicos (
                        numero_relatorio, cliente_id, responsavel_id, data_criacao,
                        formulario_servico, tipo_servico, descricao_servico, data_recebimento,
                        condicao_encontrada, placa_identificacao, acoplamento, aspectos_rotores,
                        valvulas_acopladas, data_recebimento_equip, parafusos_pinos, superficie_vedacao,
                        engrenagens, bico_injertor, rolamentos, aspecto_oleo, data_peritagem,
                        interf_desmontagem, aspecto_rotores_aba3, aspecto_carcaca, interf_mancais,
                        galeria_hidraulica, data_desmembracao, servicos_propostos, pecas_recomendadas,
                        data_pecas, cotacao_id, tempo_trabalho_total, tempo_deslocamento_total,
                        fotos, anexos_aba1, anexos_aba2, anexos_aba3, anexos_aba4
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, dados_relatorio)
                
                relatorio_id = c.lastrowid
                self.current_relatorio_id = relatorio_id
            
            # Inserir eventos dos técnicos
            for tecnico_id, tecnico_data in self.tecnicos_eventos.items():
                tree = tecnico_data['tree']
                for item in tree.get_children():
                    data_hora, tipo, evento = tree.item(item)['values']
                    c.execute("""
                        INSERT INTO eventos_campo (relatorio_id, tecnico_id, data_hora, evento, tipo)
                        VALUES (?, ?, ?, ?, ?)
                    """, (relatorio_id, tecnico_id, data_hora, evento, tipo))
            
            conn.commit()
            self.show_success("Relatório salvo com sucesso!")
            
            # Emitir evento para atualizar outros módulos
            self.emit_event('relatorio_created')
            
            # Recarregar lista
            self.carregar_relatorios()
            
        except sqlite3.Error as e:
            self.show_error(f"Erro ao salvar relatório: {e}")
            import traceback
            print(f"Erro completo: {traceback.format_exc()}")
        finally:
            conn.close()
            
    def carregar_relatorios(self):
        """Carregar lista de relatórios"""
        # Limpar lista atual
        for item in self.relatorios_tree.get_children():
            self.relatorios_tree.delete(item)
            
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        
        try:
            c.execute("""
                SELECT r.id, r.numero_relatorio, cl.nome, r.data_criacao, 
                       u.nome_completo, r.tipo_servico
                FROM relatorios_tecnicos r
                JOIN clientes cl ON r.cliente_id = cl.id
                JOIN usuarios u ON r.responsavel_id = u.id
                ORDER BY r.created_at DESC
            """)
            
            for row in c.fetchall():
                relatorio_id, numero, cliente, data, responsavel, tipo = row
                self.relatorios_tree.insert("", "end", values=(
                    numero,
                    cliente,
                    format_date(data),
                    responsavel,
                    tipo or ""
                ), tags=(relatorio_id,))
                
        except sqlite3.Error as e:
            self.show_error(f"Erro ao carregar relatórios: {e}")
        finally:
            conn.close()
            
    def buscar_relatorios(self):
        """Buscar relatórios com filtro"""
        termo = self.search_var.get().strip()
        
        # Limpar lista atual
        for item in self.relatorios_tree.get_children():
            self.relatorios_tree.delete(item)
            
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        
        try:
            if termo:
                c.execute("""
                    SELECT r.id, r.numero_relatorio, cl.nome, r.data_criacao, 
                           u.nome_completo, r.tipo_servico
                    FROM relatorios_tecnicos r
                    JOIN clientes cl ON r.cliente_id = cl.id
                    JOIN usuarios u ON r.responsavel_id = u.id
                    WHERE r.numero_relatorio LIKE ? OR cl.nome LIKE ?
                    ORDER BY r.created_at DESC
                """, (f"%{termo}%", f"%{termo}%"))
            else:
                c.execute("""
                    SELECT r.id, r.numero_relatorio, cl.nome, r.data_criacao, 
                           u.nome_completo, r.tipo_servico
                    FROM relatorios_tecnicos r
                    JOIN clientes cl ON r.cliente_id = cl.id
                    JOIN usuarios u ON r.responsavel_id = u.id
                    ORDER BY r.created_at DESC
                """)
            
            for row in c.fetchall():
                relatorio_id, numero, cliente, data, responsavel, tipo = row
                self.relatorios_tree.insert("", "end", values=(
                    numero,
                    cliente,
                    format_date(data),
                    responsavel,
                    tipo or ""
                ), tags=(relatorio_id,))
                
        except sqlite3.Error as e:
            self.show_error(f"Erro ao buscar relatórios: {e}")
        finally:
            conn.close()
            
    def editar_relatorio(self):
        """Editar relatório selecionado"""
        selected = self.relatorios_tree.selection()
        if not selected:
            self.show_warning("Selecione um relatório para editar.")
            return
            
        # Obter ID do relatório
        tags = self.relatorios_tree.item(selected[0])['tags']
        if not tags:
            return
            
        relatorio_id = tags[0]
        
        # Verificar se deve abrir no editor PDF ou no formulário
        resposta = messagebox.askyesno(
            "Edição de Relatório", 
            "Como deseja editar este relatório?\n\n"
            "Sim: Editor PDF (recomendado)\n"
            "Não: Formulário tradicional"
        )
        
        if resposta:
            # Abrir no editor PDF
            self.abrir_relatorio_editor_pdf(relatorio_id)
        else:
            # Carregar no formulário tradicional
            self.carregar_relatorio_para_edicao(relatorio_id)
            # Mudar para aba de novo relatório
            self.notebook.select(0)
        
    def carregar_relatorio_para_edicao(self, relatorio_id):
        """Carregar dados do relatório para edição"""
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        
        try:
            # Carregar dados do relatório
            c.execute("SELECT * FROM relatorios_tecnicos WHERE id = ?", (relatorio_id,))
            relatorio = c.fetchone()
            
            if not relatorio:
                self.show_error("Relatório não encontrado.")
                return
                
            # Primeiro, limpar formulário (mas preservar anexos se for edição)
            self.limpar_formulario_edicao()
            
            # Preencher campos básicos
            self.current_relatorio_id = relatorio_id
            self.numero_relatorio_var.set(relatorio[1] or "")  # numero_relatorio
            
            # Encontrar cliente no combo
            for key, value in self.clientes_dict.items():
                if value == relatorio[2]:  # cliente_id
                    self.cliente_var.set(key)
                    break
                    
            self.data_criacao_var.set(format_date(relatorio[4]) if relatorio[4] else "")
            self.formulario_servico_var.set(relatorio[5] or "")
            self.tipo_servico_var.set(relatorio[6] or "")
            
            # Descrição do serviço
            if relatorio[7]:
                self.descricao_text.insert("1.0", relatorio[7])
                
            self.data_recebimento_var.set(relatorio[8] or "")
            
            # Carregar dados das abas (índices 9-30)
            # Aba 1 (índices 9-14)
            aba1_campos = ["Cond. Encontrada", "Placa/N.Série", "Acoplamento", "Aspectos Rotores", "Válvulas Acopladas", "Data Recebimento"]
            for i, campo in enumerate(aba1_campos):
                if campo in self.aba1_vars:
                    self.aba1_vars[campo].set(relatorio[9 + i] or "")
            
            # Aba 2 (índices 15-21)
            aba2_campos = ["Parafusos/Pinos", "Superfície Vedação", "Engrenagens", "Bico Injertor", "Rolamentos", "Aspecto Óleo", "Data"]
            for i, campo in enumerate(aba2_campos):
                if campo in self.aba2_vars:
                    self.aba2_vars[campo].set(relatorio[15 + i] or "")
            
            # Aba 3 (índices 22-27)
            aba3_campos = ["Interf. Desmontagem", "Aspecto Rotores", "Aspecto Carcaça", "Interf. Mancais", "Galeria Hidráulica", "Data Desmembração"]
            for i, campo in enumerate(aba3_campos):
                if campo in self.aba3_vars:
                    self.aba3_vars[campo].set(relatorio[22 + i] or "")
            
            # Aba 4 (índices 28-30)
            if relatorio[28]:  # servicos_propostos
                self.servicos_text.insert("1.0", relatorio[28])
            if relatorio[29]:  # pecas_recomendadas
                self.pecas_text.insert("1.0", relatorio[29])
            self.data_pecas_var.set(relatorio[30] or "")
            
            # Cotação vinculada
            if relatorio[31]:  # cotacao_id
                for key, value in self.cotacoes_dict.items():
                    if value == relatorio[31]:
                        self.cotacao_var.set(key)
                        break
            
            # Carregar anexos (índices 35-38)
            for aba_num in range(1, 5):
                # Limpar anexos e listbox desta aba primeiro
                self.anexos_aba[aba_num] = []
                listbox = getattr(self, f'anexos_listbox_aba{aba_num}')
                listbox.delete(0, tk.END)
                
                anexos_str = relatorio[34 + aba_num]  # anexos_aba1, anexos_aba2, etc.
                if anexos_str:
                    try:
                        # Tentar carregar como JSON
                        self.anexos_aba[aba_num] = json.loads(anexos_str)
                    except (json.JSONDecodeError, TypeError):
                        # Fallback para formato antigo (separado por ;)
                        anexos_list = anexos_str.split(';')
                        self.anexos_aba[aba_num] = [anexo for anexo in anexos_list if anexo]
                    
                    # Atualizar listbox
                    for anexo in self.anexos_aba[aba_num]:
                        # Se for dict, usar nome; se for string, usar o nome do arquivo
                        if isinstance(anexo, dict):
                            nome_anexo = anexo.get('nome', anexo.get('path', 'Arquivo sem nome'))
                            if isinstance(nome_anexo, str) and '/' in nome_anexo:
                                nome_anexo = nome_anexo.split('/')[-1]
                        else:
                            nome_anexo = str(anexo).split('/')[-1] if isinstance(anexo, str) else str(anexo)
                        listbox.insert(tk.END, nome_anexo)
            
            # Carregar eventos dos técnicos
            self.carregar_eventos_relatorio(relatorio_id)
            
        except sqlite3.Error as e:
            self.show_error(f"Erro ao carregar relatório: {e}")
            import traceback
            print(f"Erro completo: {traceback.format_exc()}")
        finally:
            conn.close()
            
    def abrir_relatorio_editor_pdf(self, relatorio_id):
        """Abrir relatório no editor PDF para edição"""
        try:
            # Garantir que o editor PDF está carregado
            editor_module = self.main_window.editor_avancado_module
            if editor_module is None:
                # Carregar o editor se não estiver carregado
                self.main_window.load_pdf_editor()
                editor_module = self.main_window.editor_avancado_module
                
            if editor_module is None:
                self.show_error("Erro ao carregar o editor PDF.")
                return
            
            # Gerar PDF do relatório atual
            from pdf_generators.relatorio_tecnico import gerar_relatorio_pdf
            
            # Obter dados do relatório
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            c.execute("SELECT * FROM relatorios_tecnicos WHERE id = ?", (relatorio_id,))
            relatorio = c.fetchone()
            
            if not relatorio:
                self.show_error("Relatório não encontrado.")
                conn.close()
                return
            
            # Gerar PDF temporário
            import tempfile
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
                pdf_path = temp_file.name
            
            success = gerar_relatorio_pdf(relatorio_id, pdf_path)
            conn.close()
            
            if success:
                # Carregar PDF no editor
                editor_module.load_pdf_for_editing(pdf_path, relatorio_id)
                
                # Mudar para aba do editor PDF
                notebook = self.main_window.notebook
                for i in range(notebook.index("end")):
                    if "Editor Avançado" in notebook.tab(i, "text"):
                        notebook.select(i)
                        break
                        
                self.show_info("Relatório aberto no Editor PDF para edição.")
            else:
                self.show_error("Erro ao gerar PDF do relatório.")
                
        except Exception as e:
            self.show_error(f"Erro ao abrir relatório no editor PDF: {e}")
            import traceback
            print(f"Erro completo: {traceback.format_exc()}")

    def carregar_eventos_relatorio(self, relatorio_id):
        """Carregar eventos dos técnicos do relatório"""
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        
        try:
            c.execute("""
                SELECT ec.tecnico_id, u.nome_completo, ec.data_hora, ec.evento, ec.tipo
                FROM eventos_campo ec
                JOIN usuarios u ON ec.tecnico_id = u.id
                WHERE ec.relatorio_id = ?
                ORDER BY ec.tecnico_id, ec.data_hora
            """, (relatorio_id,))
            
            eventos = c.fetchall()
            tecnicos_adicionados = set()
            
            for evento in eventos:
                tecnico_id, tecnico_nome, data_hora, descricao, tipo = evento
                
                # Adicionar técnico se ainda não foi adicionado
                if tecnico_id not in tecnicos_adicionados:
                    # Simular seleção do técnico
                    for key, value in self.tecnicos_dict.items():
                        if value == tecnico_id:
                            self.tecnico_var.set(key)
                            self.adicionar_tecnico()
                            break
                    tecnicos_adicionados.add(tecnico_id)
                
                # Adicionar evento
                if tecnico_id in self.tecnicos_eventos:
                    tree = self.tecnicos_eventos[tecnico_id]['tree']
                    tree.insert("", "end", values=(data_hora, tipo, descricao))
                    
        except sqlite3.Error as e:
            self.show_error(f"Erro ao carregar eventos: {e}")
        finally:
            conn.close()
            
    def duplicar_relatorio(self):
        """Duplicar relatório selecionado"""
        selected = self.relatorios_tree.selection()
        if not selected:
            self.show_warning("Selecione um relatório para duplicar.")
            return
            
        # Obter ID do relatório
        tags = self.relatorios_tree.item(selected[0])['tags']
        if not tags:
            return
            
        relatorio_id = tags[0]
        self.carregar_relatorio_para_edicao(relatorio_id)
        
        # Limpar ID e gerar novo número
        self.current_relatorio_id = None
        numero = f"REL-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        self.numero_relatorio_var.set(numero)
        self.data_criacao_var.set(datetime.now().strftime('%d/%m/%Y'))
        
        # Mudar para aba de novo relatório
        self.notebook.select(0)
        
    def gerar_pdf(self):
        """Gerar PDF do relatório atual"""
        if not self.current_relatorio_id:
            self.show_warning("Salve o relatório antes de gerar o PDF.")
            return
            
        sucesso, resultado = gerar_pdf_relatorio(self.current_relatorio_id, DB_NAME)
        
        if sucesso:
            self.show_success(f"PDF gerado com sucesso!\nLocal: {resultado}")
        else:
            self.show_error(f"Erro ao gerar PDF: {resultado}")
            
    def gerar_pdf_selecionado(self):
        """Gerar PDF do relatório selecionado"""
        selected = self.relatorios_tree.selection()
        if not selected:
            self.show_warning("Selecione um relatório para gerar PDF.")
            return
            
        # Obter ID do relatório
        tags = self.relatorios_tree.item(selected[0])['tags']
        if not tags:
            return
            
        relatorio_id = tags[0]
        sucesso, resultado = gerar_pdf_relatorio(relatorio_id, DB_NAME)
        
        if sucesso:
            self.show_success(f"PDF gerado com sucesso!\nLocal: {resultado}")
        else:
            self.show_error(f"Erro ao gerar PDF: {resultado}")
            
    def handle_event(self, event_type, data=None):
        """Manipular eventos do sistema"""
        if event_type == 'cliente_created':
            self.refresh_clientes()
            print("Lista de clientes atualizada automaticamente no relatório!")
        elif event_type == 'tecnico_created':
            self.refresh_tecnicos()
            print("Lista de técnicos atualizada automaticamente!")

    def create_relatorios_gerais_tab(self):
        """Criar aba de Relatórios Gerais com extração de dados personalizados"""
        relatorios_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(relatorios_frame, text="Relatórios Gerais")
        
        container = tk.Frame(relatorios_frame, bg='white', padx=20, pady=20)
        container.pack(fill="both", expand=True)
        
        # Header da seção
        header_frame = tk.Frame(container, bg='white')
        header_frame.pack(fill="x", pady=(0, 20))
        
        title_label = tk.Label(header_frame, text="Relatórios Gerais - Extração de Dados", 
                               font=('Arial', 16, 'bold'), bg='white', fg='#1e293b')
        title_label.pack(side="left")
        
        # Frame principal dividido em duas colunas
        main_frame = tk.Frame(container, bg='white')
        main_frame.pack(fill="both", expand=True)
        
        # Coluna esquerda - Filtros e Configurações (40%)
        left_frame = tk.Frame(main_frame, bg='white')
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        # Coluna direita - Resultados (60%)
        right_frame = tk.Frame(main_frame, bg='white')
        right_frame.pack(side="right", fill="both", expand=True, padx=(10, 0))
        
        # NOVO: Coluna de indicadores (cards) - largura aumentada
        indicators_frame = tk.Frame(main_frame, bg='#f8fafc', width=320)
        indicators_frame.pack(side="right", fill="y", padx=(10, 0))
        self.create_indicators_cards(indicators_frame)
        # Fim dos indicadores
        
        # Seção de Filtros
        self.create_filtros_section(left_frame)
        
        # Seção de Tipos de Relatório
        self.create_tipos_relatorio_section(left_frame)
        
        # Seção de Resultados
        self.create_resultados_section(right_frame)
        
    def create_indicators_cards(self, parent):
        """Cria cards de indicadores na lateral direita da aba Relatórios Gerais"""
        # Card 1: Total de Relatórios
        self.create_stat_card(parent, "📋 Relatórios", "0", "#3b82f6", "Este mês: 0", big=True).pack(fill="x", pady=(0, 16))
        # Card 2: Faturamento Total
        self.create_stat_card(parent, "💵 Faturamento", "R$ 0,00", "#10b981", "Média: R$ 0,00", big=True).pack(fill="x", pady=(0, 16))
        # Card 3: Performance
        self.create_stat_card(parent, "📈 Performance", "0%", "#ef4444", "Taxa de Aprovação", big=True).pack(fill="x", pady=(0, 16))
    
    def create_stat_card(self, parent, title, value, color, subtitle="", big=False):
        card = tk.Frame(parent, bg='white', relief='solid', bd=1)
        pad_x = 32 if big else 20
        pad_y = 28 if big else 15
        font_title = ('Arial', 18, 'bold') if big else ('Arial', 14, 'bold')
        font_value = ('Arial', 38, 'bold') if big else ('Arial', 28, 'bold')
        font_subtitle = ('Arial', 12) if big else ('Arial', 10)
        inner_frame = tk.Frame(card, bg='white', padx=pad_x, pady=pad_y)
        inner_frame.pack(fill="both", expand=True)
        title_label = tk.Label(inner_frame, text=title, font=font_title, bg='white', fg=color)
        title_label.pack(anchor="w")
        value_label = tk.Label(inner_frame, text=value, font=font_value, bg='white', fg='#1e293b')
        value_label.pack(anchor="w", pady=(5, 0))
        if subtitle:
            subtitle_label = tk.Label(inner_frame, text=subtitle, font=font_subtitle, bg='white', fg='#64748b')
            subtitle_label.pack(anchor="w", pady=(2, 0))
        return card
    
    def create_filtros_section(self, parent):
        """Criar seção de filtros para relatórios gerais"""
        section_frame = self.create_section_frame(parent, "Filtros de Data")
        section_frame.pack(fill="x", pady=(0, 15))
        
        fields_frame = tk.Frame(section_frame, bg='white')
        fields_frame.pack(fill="x", padx=10, pady=10)
        
        # Variáveis de filtro
        self.data_inicio_var = tk.StringVar()
        self.data_fim_var = tk.StringVar()
        self.usuario_filtro_var = tk.StringVar(value="Todos")
        self.cliente_filtro_var = tk.StringVar(value="Todos")
        
        # Data de início
        tk.Label(fields_frame, text="Data Início:", font=('Arial', 10, 'bold'), bg='white').grid(row=0, column=0, sticky="w", pady=5)
        tk.Entry(fields_frame, textvariable=self.data_inicio_var, font=('Arial', 10), width=15).grid(row=0, column=1, sticky="ew", padx=(10, 0), pady=5)
        
        # Data fim
        tk.Label(fields_frame, text="Data Fim:", font=('Arial', 10, 'bold'), bg='white').grid(row=1, column=0, sticky="w", pady=5)
        tk.Entry(fields_frame, textvariable=self.data_fim_var, font=('Arial', 10), width=15).grid(row=1, column=1, sticky="ew", padx=(10, 0), pady=5)
        
        # Filtro por usuário
        tk.Label(fields_frame, text="Usuário:", font=('Arial', 10, 'bold'), bg='white').grid(row=2, column=0, sticky="w", pady=5)
        self.usuario_combo = ttk.Combobox(fields_frame, textvariable=self.usuario_filtro_var, width=20)
        self.usuario_combo.grid(row=2, column=1, sticky="ew", padx=(10, 0), pady=5)
        
        # Filtro por cliente
        tk.Label(fields_frame, text="Cliente:", font=('Arial', 10, 'bold'), bg='white').grid(row=3, column=0, sticky="w", pady=5)
        self.cliente_combo = ttk.Combobox(fields_frame, textvariable=self.cliente_filtro_var, width=20)
        self.cliente_combo.grid(row=3, column=1, sticky="ew", padx=(10, 0), pady=5)
        
        # Carregar dados para os combos
        self.carregar_filtros_dados()
        
        fields_frame.grid_columnconfigure(1, weight=1)
        
    def create_tipos_relatorio_section(self, parent):
        """Criar seção de tipos de relatório"""
        section_frame = self.create_section_frame(parent, "Tipos de Relatório")
        section_frame.pack(fill="both", expand=True, pady=(15, 0))
        
        types_frame = tk.Frame(section_frame, bg='white')
        types_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Variável para tipo de relatório
        self.tipo_relatorio_var = tk.StringVar(value="cotacoes_status")
        
        # Tipos de relatório disponíveis
        tipos = [
            ("cotacoes_status", "Cotações por Status"),
            ("cotacoes_usuario", "Cotações por Usuário"),
            ("faturamento_usuario", "Faturamento por Usuário"),
            ("faturamento_produto", "Faturamento por Tipo de Produto"),
            ("propostas_usuario", "Quantidade de Propostas por Usuário"),
            ("relatorios_tecnicos", "Relatórios Técnicos por Período"),
            ("clientes_atividade", "Atividade por Cliente"),
            ("performance_usuarios", "Performance dos Usuários")
        ]
        
        for i, (value, text) in enumerate(tipos):
            rb = tk.Radiobutton(types_frame, text=text, variable=self.tipo_relatorio_var, value=value,
                               font=('Arial', 10), bg='white', anchor='w')
            rb.grid(row=i, column=0, sticky="ew", pady=2)
        
        types_frame.grid_columnconfigure(0, weight=1)
        
        # Botão para gerar relatório
        gerar_btn = self.create_button(types_frame, "📊 Gerar Relatório", self.gerar_relatorio_geral, bg='#10b981')
        gerar_btn.grid(row=len(tipos), column=0, pady=(20, 0), sticky="ew")
        
    def create_resultados_section(self, parent):
        """Criar seção de resultados"""
        section_frame = self.create_section_frame(parent, "Resultados")
        section_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        # Frame para treeview e scrollbar
        tree_frame = tk.Frame(section_frame, bg='white')
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Treeview para resultados
        self.resultados_tree = ttk.Treeview(tree_frame, show="headings", height=15)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.resultados_tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.resultados_tree.xview)
        
        self.resultados_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Grid layout
        self.resultados_tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Frame para botões de exportação
        export_frame = tk.Frame(section_frame, bg='white')
        export_frame.pack(fill="x", padx=10, pady=(10, 0))
        
        # Botões de exportação
        export_csv_btn = self.create_button(export_frame, "📄 Exportar CSV", self.exportar_csv, bg='#3b82f6')
        export_csv_btn.pack(side="left", padx=(0, 10))
        
        export_excel_btn = self.create_button(export_frame, "📊 Exportar Excel", self.exportar_excel, bg='#10b981')
        export_excel_btn.pack(side="left")
        
    def carregar_filtros_dados(self):
        """Carregar dados para os filtros"""
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        
        try:
            # Carregar usuários
            c.execute("SELECT nome_completo FROM usuarios WHERE nome_completo IS NOT NULL ORDER BY nome_completo")
            usuarios = ["Todos"] + [row[0] for row in c.fetchall()]
            self.usuario_combo['values'] = usuarios
            
            # Carregar clientes
            c.execute("SELECT nome FROM clientes ORDER BY nome")
            clientes = ["Todos"] + [row[0] for row in c.fetchall()]
            self.cliente_combo['values'] = clientes
            
        except sqlite3.Error as e:
            print(f"Erro ao carregar dados dos filtros: {e}")
        finally:
            conn.close()
            
    def gerar_relatorio_geral(self):
        """Gerar relatório geral baseado no tipo selecionado"""
        tipo = self.tipo_relatorio_var.get()
        
        # Limpar resultados anteriores
        for item in self.resultados_tree.get_children():
            self.resultados_tree.delete(item)
        
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        
        try:
            if tipo == "cotacoes_status":
                self.gerar_relatorio_cotacoes_status(c)
            elif tipo == "cotacoes_usuario":
                self.gerar_relatorio_cotacoes_usuario(c)
            elif tipo == "faturamento_usuario":
                self.gerar_relatorio_faturamento_usuario(c)
            elif tipo == "faturamento_produto":
                self.gerar_relatorio_faturamento_produto(c)
            elif tipo == "propostas_usuario":
                self.gerar_relatorio_propostas_usuario(c)
            elif tipo == "relatorios_tecnicos":
                self.gerar_relatorio_relatorios_tecnicos(c)
            elif tipo == "clientes_atividade":
                self.gerar_relatorio_clientes_atividade(c)
            elif tipo == "performance_usuarios":
                self.gerar_relatorio_performance_usuarios(c)
                
        except sqlite3.Error as e:
            self.show_error(f"Erro ao gerar relatório: {e}")
        finally:
            conn.close()
            
    def gerar_relatorio_cotacoes_status(self, cursor):
        """Gerar relatório de cotações por status"""
        # Configurar colunas
        columns = ("status", "quantidade", "valor_total", "percentual")
        self.resultados_tree["columns"] = columns
        
        for col in columns:
            self.resultados_tree.heading(col, text=col.replace("_", " ").title())
            self.resultados_tree.column(col, width=150)
        
        # Buscar dados
        cursor.execute("""
            SELECT status, COUNT(*) as quantidade, COALESCE(SUM(valor_total), 0) as valor_total
            FROM cotacoes 
            GROUP BY status
            ORDER BY quantidade DESC
        """)
        
        resultados = cursor.fetchall()
        total_cotacoes = sum(row[1] for row in resultados)
        
        # Inserir dados
        for status, quantidade, valor_total in resultados:
            percentual = (quantidade / total_cotacoes * 100) if total_cotacoes > 0 else 0
            self.resultados_tree.insert("", "end", values=(
                status,
                quantidade,
                f"R$ {valor_total:,.2f}",
                f"{percentual:.1f}%"
            ))
            
    def gerar_relatorio_cotacoes_usuario(self, cursor):
        """Gerar relatório de cotações por usuário"""
        columns = ("usuario", "quantidade", "valor_total", "media_valor")
        self.resultados_tree["columns"] = columns
        
        for col in columns:
            self.resultados_tree.heading(col, text=col.replace("_", " ").title())
            self.resultados_tree.column(col, width=150)
        
        cursor.execute("""
            SELECT u.nome_completo, COUNT(c.id) as quantidade, 
                   COALESCE(SUM(c.valor_total), 0) as valor_total,
                   COALESCE(AVG(c.valor_total), 0) as media_valor
            FROM usuarios u
            LEFT JOIN cotacoes c ON u.id = c.responsavel_id
            GROUP BY u.id, u.nome_completo
            ORDER BY quantidade DESC
        """)
        
        for nome, quantidade, valor_total, media_valor in cursor.fetchall():
            self.resultados_tree.insert("", "end", values=(
                nome or "Usuário sem nome",
                quantidade,
                f"R$ {valor_total:,.2f}",
                f"R$ {media_valor:,.2f}"
            ))
            
    def gerar_relatorio_faturamento_usuario(self, cursor):
        """Gerar relatório de faturamento por usuário"""
        columns = ("usuario", "cotacoes_aprovadas", "valor_faturado", "ticket_medio")
        self.resultados_tree["columns"] = columns
        
        for col in columns:
            self.resultados_tree.heading(col, text=col.replace("_", " ").title())
            self.resultados_tree.column(col, width=150)
        
        cursor.execute("""
            SELECT u.nome_completo, COUNT(c.id) as cotacoes_aprovadas, 
                   COALESCE(SUM(c.valor_total), 0) as valor_faturado,
                   COALESCE(AVG(c.valor_total), 0) as ticket_medio
            FROM usuarios u
            LEFT JOIN cotacoes c ON u.id = c.responsavel_id AND c.status = 'Aprovada'
            GROUP BY u.id, u.nome_completo
            ORDER BY valor_faturado DESC
        """)
        
        for nome, cotacoes, valor, ticket in cursor.fetchall():
            self.resultados_tree.insert("", "end", values=(
                nome or "Usuário sem nome",
                cotacoes,
                f"R$ {valor:,.2f}",
                f"R$ {ticket:,.2f}"
            ))
            
    def gerar_relatorio_faturamento_produto(self, cursor):
        """Gerar relatório de faturamento por tipo de produto"""
        columns = ("tipo_produto", "quantidade_vendida", "valor_total", "participacao")
        self.resultados_tree["columns"] = columns
        
        for col in columns:
            self.resultados_tree.heading(col, text=col.replace("_", " ").title())
            self.resultados_tree.column(col, width=150)
        
        cursor.execute("""
            SELECT ic.tipo, SUM(ic.quantidade) as quantidade, 
                   SUM(ic.valor_total_item) as valor_total
            FROM itens_cotacao ic
            JOIN cotacoes c ON ic.cotacao_id = c.id
            WHERE c.status = 'Aprovada'
            GROUP BY ic.tipo
            ORDER BY valor_total DESC
        """)
        
        resultados = cursor.fetchall()
        total_valor = sum(row[2] for row in resultados)
        
        for tipo, quantidade, valor in resultados:
            participacao = (valor / total_valor * 100) if total_valor > 0 else 0
            self.resultados_tree.insert("", "end", values=(
                tipo,
                f"{quantidade:.0f}",
                f"R$ {valor:,.2f}",
                f"{participacao:.1f}%"
            ))
            
    def gerar_relatorio_propostas_usuario(self, cursor):
        """Gerar relatório de quantidade de propostas por usuário"""
        columns = ("usuario", "total_propostas", "aprovadas", "rejeitadas", "em_aberto", "taxa_aprovacao")
        self.resultados_tree["columns"] = columns
        
        for col in columns:
            self.resultados_tree.heading(col, text=col.replace("_", " ").title())
            self.resultados_tree.column(col, width=120)
        
        cursor.execute("""
            SELECT u.nome_completo,
                   COUNT(c.id) as total,
                   SUM(CASE WHEN c.status = 'Aprovada' THEN 1 ELSE 0 END) as aprovadas,
                   SUM(CASE WHEN c.status = 'Rejeitada' THEN 1 ELSE 0 END) as rejeitadas,
                   SUM(CASE WHEN c.status = 'Em Aberto' THEN 1 ELSE 0 END) as em_aberto
            FROM usuarios u
            LEFT JOIN cotacoes c ON u.id = c.responsavel_id
            GROUP BY u.id, u.nome_completo
            ORDER BY total DESC
        """)
        
        for nome, total, aprovadas, rejeitadas, em_aberto in cursor.fetchall():
            taxa = (aprovadas / total * 100) if total > 0 else 0
            self.resultados_tree.insert("", "end", values=(
                nome or "Usuário sem nome",
                total,
                aprovadas,
                rejeitadas,
                em_aberto,
                f"{taxa:.1f}%"
            ))
            
    def gerar_relatorio_relatorios_tecnicos(self, cursor):
        """Gerar relatório de relatórios técnicos por período"""
        columns = ("cliente", "numero_relatorio", "data_criacao", "responsavel", "tipo_servico")
        self.resultados_tree["columns"] = columns
        
        for col in columns:
            self.resultados_tree.heading(col, text=col.replace("_", " ").title())
            self.resultados_tree.column(col, width=150)
        
        cursor.execute("""
            SELECT cl.nome, rt.numero_relatorio, rt.data_criacao, 
                   u.nome_completo, rt.tipo_servico
            FROM relatorios_tecnicos rt
            JOIN clientes cl ON rt.cliente_id = cl.id
            JOIN usuarios u ON rt.responsavel_id = u.id
            ORDER BY rt.data_criacao DESC
        """)
        
        for cliente, numero, data, responsavel, tipo in cursor.fetchall():
            self.resultados_tree.insert("", "end", values=(
                cliente,
                numero,
                data,
                responsavel or "Não informado",
                tipo or "Não informado"
            ))
            
    def gerar_relatorio_clientes_atividade(self, cursor):
        """Gerar relatório de atividade por cliente"""
        columns = ("cliente", "cotacoes", "relatorios", "ultima_atividade", "valor_total")
        self.resultados_tree["columns"] = columns
        
        for col in columns:
            self.resultados_tree.heading(col, text=col.replace("_", " ").title())
            self.resultados_tree.column(col, width=150)
        
        cursor.execute("""
            SELECT cl.nome,
                   COUNT(DISTINCT c.id) as cotacoes,
                   COUNT(DISTINCT rt.id) as relatorios,
                   MAX(COALESCE(c.data_criacao, rt.data_criacao)) as ultima_atividade,
                   COALESCE(SUM(c.valor_total), 0) as valor_total
            FROM clientes cl
            LEFT JOIN cotacoes c ON cl.id = c.cliente_id
            LEFT JOIN relatorios_tecnicos rt ON cl.id = rt.cliente_id
            GROUP BY cl.id, cl.nome
            ORDER BY ultima_atividade DESC
        """)
        
        for nome, cotacoes, relatorios, ultima, valor in cursor.fetchall():
            self.resultados_tree.insert("", "end", values=(
                nome,
                cotacoes,
                relatorios,
                ultima or "Nunca",
                f"R$ {valor:,.2f}"
            ))
            
    def gerar_relatorio_performance_usuarios(self, cursor):
        """Gerar relatório de performance dos usuários"""
        columns = ("usuario", "cotacoes_total", "aprovadas", "valor_medio", "eficiencia")
        self.resultados_tree["columns"] = columns
        
        for col in columns:
            self.resultados_tree.heading(col, text=col.replace("_", " ").title())
            self.resultados_tree.column(col, width=150)
        
        cursor.execute("""
            SELECT u.nome_completo,
                   COUNT(c.id) as total,
                   SUM(CASE WHEN c.status = 'Aprovada' THEN 1 ELSE 0 END) as aprovadas,
                   AVG(CASE WHEN c.status = 'Aprovada' THEN c.valor_total ELSE NULL END) as valor_medio,
                   CASE 
                       WHEN COUNT(c.id) > 0 THEN 
                           CAST(SUM(CASE WHEN c.status = 'Aprovada' THEN 1 ELSE 0 END) AS FLOAT) / COUNT(c.id) * 100
                       ELSE 0 
                   END as eficiencia
            FROM usuarios u
            LEFT JOIN cotacoes c ON u.id = c.responsavel_id
            GROUP BY u.id, u.nome_completo
            HAVING COUNT(c.id) > 0
            ORDER BY eficiencia DESC
        """)
        
        for nome, total, aprovadas, valor_medio, eficiencia in cursor.fetchall():
            self.resultados_tree.insert("", "end", values=(
                nome or "Usuário sem nome",
                total,
                aprovadas,
                f"R$ {valor_medio or 0:,.2f}",
                f"{eficiencia:.1f}%"
            ))
            
    def exportar_csv(self):
        """Exportar resultados para CSV"""
        try:
            import csv
            from tkinter import filedialog
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )
            
            if filename:
                with open(filename, 'w', newline='', encoding='utf-8') as file:
                    writer = csv.writer(file)
                    
                    # Cabeçalhos
                    headers = [self.resultados_tree.heading(col)['text'] for col in self.resultados_tree['columns']]
                    writer.writerow(headers)
                    
                    # Dados
                    for item in self.resultados_tree.get_children():
                        values = self.resultados_tree.item(item)['values']
                        writer.writerow(values)
                
                self.show_success(f"Relatório exportado para: {filename}")
                
        except Exception as e:
            self.show_error(f"Erro ao exportar CSV: {e}")
            
    def exportar_excel(self):
        """Exportar resultados para Excel"""
        try:
            # Tentar importar openpyxl
            try:
                from openpyxl import Workbook
            except ImportError:
                self.show_warning("Para exportar Excel, instale: pip install openpyxl")
                return
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if filename:
                wb = Workbook()
                ws = wb.active
                ws.title = "Relatório"
                
                # Cabeçalhos
                headers = [self.resultados_tree.heading(col)['text'] for col in self.resultados_tree['columns']]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Dados
                for row, item in enumerate(self.resultados_tree.get_children(), 2):
                    values = self.resultados_tree.item(item)['values']
                    for col, value in enumerate(values, 1):
                        ws.cell(row=row, column=col, value=value)
                
                wb.save(filename)
                self.show_success(f"Relatório exportado para: {filename}")
                
        except Exception as e:
            self.show_error(f"Erro ao exportar Excel: {e}")